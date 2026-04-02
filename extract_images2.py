import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import requests
import re

# 创建输出目录
output_dir = 'd:/code/web/public/assets/products'
os.makedirs(output_dir, exist_ok=True)

# 使用 openpyxl 读取以获取超链接和图片信息
wb = openpyxl.load_workbook('d:/code/web/src/Amazon商品数据（含图片）.xlsx')
ws = wb.active

print(f"工作表名称: {ws.title}")
print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

# 检查每一行的超链接
image_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # 第一列通常是主图
    cell = row[0]  # 第一列
    
    # 检查单元格是否有超链接
    if cell.hyperlink:
        link = cell.hyperlink.target
        print(f"行 {row_idx}: 发现超链接 - {link}")
        
        # 如果是图片链接，尝试下载
        if link and ('.jpg' in link or '.png' in link or '.jpeg' in link or 'image' in link):
            try:
                response = requests.get(link, timeout=10)
                if response.status_code == 200:
                    # 确定文件扩展名
                    if '.jpg' in link or '.jpeg' in link:
                        ext = 'jpg'
                    elif '.png' in link:
                        ext = 'png'
                    else:
                        ext = 'jpg'
                    
                    filename = f"amazon_product_{row_idx-1:03d}.{ext}"
                    filepath = os.path.join(output_dir, filename)
                    
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    
                    image_count += 1
                    print(f"  已下载: {filename}")
            except Exception as e:
                print(f"  下载失败: {e}")
    
    # 检查单元格值是否是 URL
    cell_value = cell.value
    if cell_value and isinstance(cell_value, str):
        if cell_value.startswith('http') and ('.jpg' in cell_value or '.png' in cell_value or 'amazon' in cell_value):
            print(f"行 {row_idx}: 发现 URL - {cell_value[:80]}...")
            try:
                response = requests.get(cell_value, timeout=10)
                if response.status_code == 200:
                    ext = 'jpg' if '.jpg' in cell_value else 'png'
                    filename = f"amazon_product_{row_idx-1:03d}.{ext}"
                    filepath = os.path.join(output_dir, filename)
                    
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    
                    image_count += 1
                    print(f"  已下载: {filename}")
            except Exception as e:
                print(f"  下载失败: {e}")

print(f"\n总共下载了 {image_count} 张图片")
print(f"图片已保存到: {output_dir}")

# 列出目录中的文件
files = os.listdir(output_dir)
print(f"\n目录中的文件 ({len(files)} 个):")
for f in files[:20]:  # 只显示前20个
    print(f"  - {f}")
if len(files) > 20:
    print(f"  ... 还有 {len(files)-20} 个文件")
