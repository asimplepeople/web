import openpyxl
from openpyxl import load_workbook
import os
from PIL import Image
import io

# 创建输出目录
output_dir = 'd:/code/web/public/assets/products/amazon'
os.makedirs(output_dir, exist_ok=True)

# 加载工作簿
wb = load_workbook('d:/code/web/src/Amazon商品数据（含图片）(1).xlsx')

print(f"工作表列表: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*50}")
    print(f"工作表: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 检查是否有图片
    if hasattr(ws, '_images'):
        print(f"找到 {len(ws._images)} 个图片对象")
        
        for idx, img in enumerate(ws._images):
            try:
                print(f"\n处理图片 {idx + 1}:")
                print(f"  类型: {type(img)}")
                
                # 获取图片数据
                image_data = None
                if hasattr(img, '_data'):
                    image_data = img._data
                    print(f"  数据来源: _data")
                elif hasattr(img, 'ref'):
                    image_data = img.ref
                    print(f"  数据来源: ref")
                elif hasattr(img, 'blob'):
                    image_data = img.blob
                    print(f"  数据来源: blob")
                
                if image_data:
                    print(f"  数据大小: {len(image_data) if isinstance(image_data, (bytes, bytearray)) else 'N/A'} bytes")
                    
                    # 确定图片格式
                    img_format = 'png'
                    if hasattr(img, 'format'):
                        img_format = img.format.lower()
                    elif hasattr(img, 'content_type'):
                        content_type = img.content_type
                        if 'jpeg' in content_type or 'jpg' in content_type:
                            img_format = 'jpg'
                        elif 'png' in content_type:
                            img_format = 'png'
                    
                    # 尝试从数据头检测格式
                    if isinstance(image_data, (bytes, bytearray)):
                        if image_data[:2] == b'\xff\xd8':
                            img_format = 'jpg'
                        elif image_data[:8] == b'\x89PNG\r\n\x1a\n':
                            img_format = 'png'
                        elif image_data[:4] == b'GIF8':
                            img_format = 'gif'
                    
                    # 生成文件名
                    filename = f"amazon_product_{idx + 1:03d}.{img_format}"
                    filepath = os.path.join(output_dir, filename)
                    
                    # 保存图片
                    if isinstance(image_data, (bytes, bytearray)):
                        with open(filepath, 'wb') as f:
                            f.write(image_data)
                        print(f"  已保存: {filename}")
                    else:
                        print(f"  警告: 数据类型不是 bytes，尝试转换...")
                        try:
                            with open(filepath, 'wb') as f:
                                f.write(bytes(image_data))
                            print(f"  已保存: {filename}")
                        except Exception as e:
                            print(f"  保存失败: {e}")
                else:
                    print(f"  警告: 无法获取图片数据")
                    
            except Exception as e:
                print(f"  处理图片时出错: {e}")
                import traceback
                traceback.print_exc()
    else:
        print("该工作表没有 _images 属性")

# 列出保存的文件
print(f"\n{'='*50}")
print(f"图片已保存到: {output_dir}")
if os.path.exists(output_dir):
    files = os.listdir(output_dir)
    amazon_files = [f for f in files if f.startswith('amazon_product_')]
    print(f"共保存了 {len(amazon_files)} 张图片:")
    for f in sorted(amazon_files)[:20]:
        filepath = os.path.join(output_dir, f)
        size = os.path.getsize(filepath)
        print(f"  - {f} ({size} bytes)")
    if len(amazon_files) > 20:
        print(f"  ... 还有 {len(amazon_files) - 20} 张图片")
