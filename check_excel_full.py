import openpyxl
from openpyxl import load_workbook
import os

# 加载工作簿
wb = load_workbook('d:/code/web/src/Amazon商品数据（含图片）.xlsx')

print(f"工作表列表: {wb.sheetnames}")
print(f"工作表数量: {len(wb.sheetnames)}")

# 检查每个工作表
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*50}")
    print(f"工作表: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 检查是否有图片
    if hasattr(ws, '_images'):
        print(f"图片数量: {len(ws._images)}")
        for idx, img in enumerate(ws._images):
            print(f"  图片 {idx}: {type(img)}")
            if hasattr(img, 'anchor'):
                print(f"    位置: {img.anchor}")
    else:
        print("图片数量: 0 (无 _images 属性)")
    
    # 检查是否有图表
    if hasattr(ws, '_charts'):
        print(f"图表数量: {len(ws._charts)}")
    
    # 打印前几行数据
    print("\n前3行数据:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), start=1):
        print(f"  行 {row_idx}: {row}")

# 检查是否有外部链接
print(f"\n{'='*50}")
print("检查外部链接...")
if hasattr(wb, 'external_links'):
    print(f"外部链接: {wb.external_links}")

# 检查 rels (关系)
print("\n检查关系文件...")
rels_dir = os.path.join('d:/code/web/src/', '_rels')
if os.path.exists(rels_dir):
    print(f"找到 _rels 目录: {rels_dir}")
else:
    print("未找到 _rels 目录")
