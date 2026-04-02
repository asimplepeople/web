import openpyxl
from openpyxl import load_workbook
import os
import zipfile
import shutil

# 创建输出目录
output_dir = 'd:/code/web/public/assets/products/amazon'
os.makedirs(output_dir, exist_ok=True)

# Excel 文件路径
excel_path = 'd:/code/web/src/Amazon商品数据（含图片）(1).xlsx'

# 方法1: 使用 zipfile 直接解压 xlsx 文件获取图片
print("方法1: 从 xlsx 文件中提取图片...")

try:
    with zipfile.ZipFile(excel_path, 'r') as zip_ref:
        # 列出所有文件
        all_files = zip_ref.namelist()
        print(f"xlsx 文件中的内容: {len(all_files)} 个文件/目录")
        
        # 查找媒体文件
        media_files = [f for f in all_files if 'media' in f.lower()]
        print(f"找到 {len(media_files)} 个媒体相关文件")
        
        # 提取图片
        image_count = 0
        for file_path in media_files:
            if file_path.startswith('xl/media/'):
                # 获取文件名
                filename = os.path.basename(file_path)
                # 确定扩展名
                if '.' in filename:
                    ext = filename.split('.')[-1].lower()
                else:
                    ext = 'jpg'
                
                # 生成新的文件名
                new_filename = f"amazon_product_{image_count + 1:03d}.{ext}"
                output_path = os.path.join(output_dir, new_filename)
                
                # 提取文件
                with zip_ref.open(file_path) as src:
                    with open(output_path, 'wb') as dst:
                        dst.write(src.read())
                
                image_count += 1
                print(f"  已提取: {new_filename} (来自 {file_path})")
        
        print(f"\n成功提取 {image_count} 张图片")
        
except Exception as e:
    print(f"方法1失败: {e}")
    print("\n尝试方法2...")

# 方法2: 使用 openpyxl 的另一种方式
print("\n方法2: 使用 openpyxl 提取图片...")

try:
    wb = load_workbook(excel_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n处理工作表: {sheet_name}")
        
        if hasattr(ws, '_images') and ws._images:
            print(f"找到 {len(ws._images)} 个图片对象")
            
            for idx, img in enumerate(ws._images):
                try:
                    # 尝试不同的方式获取图片数据
                    image_data = None
                    
                    # 方式1: 直接访问 _data 属性（如果是属性）
                    if hasattr(img, '_data') and not callable(img._data):
                        image_data = img._data
                    # 方式2: 调用 _data() 方法
                    elif hasattr(img, '_data') and callable(img._data):
                        image_data = img._data()
                    # 方式3: 访问 ref
                    elif hasattr(img, 'ref'):
                        image_data = img.ref
                    # 方式4: 访问 blob
                    elif hasattr(img, 'blob'):
                        image_data = img.blob
                    
                    if image_data and isinstance(image_data, (bytes, bytearray)):
                        # 确定格式
                        ext = 'jpg'
                        if image_data[:2] == b'\xff\xd8':
                            ext = 'jpg'
                        elif image_data[:8] == b'\x89PNG\r\n\x1a\n':
                            ext = 'png'
                        elif image_data[:4] == b'GIF8':
                            ext = 'gif'
                        
                        filename = f"amazon_product_v2_{idx + 1:03d}.{ext}"
                        filepath = os.path.join(output_dir, filename)
                        
                        with open(filepath, 'wb') as f:
                            f.write(image_data)
                        
                        print(f"  已保存: {filename} ({len(image_data)} bytes)")
                    else:
                        print(f"  图片 {idx + 1}: 无法获取有效数据")
                        
                except Exception as e:
                    print(f"  图片 {idx + 1} 处理失败: {e}")
        else:
            print("没有找到图片对象")
            
except Exception as e:
    print(f"方法2失败: {e}")

# 列出最终结果
print(f"\n{'='*50}")
print(f"图片保存目录: {output_dir}")
if os.path.exists(output_dir):
    files = [f for f in os.listdir(output_dir) if f.startswith('amazon_product_')]
    files.sort()
    print(f"共保存了 {len(files)} 张图片:")
    for f in files:
        filepath = os.path.join(output_dir, f)
        size = os.path.getsize(filepath)
        print(f"  - {f} ({size} bytes)")
