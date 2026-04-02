import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import shutil

# 创建输出目录
output_dir = 'd:/code/web/public/assets/products'
os.makedirs(output_dir, exist_ok=True)

# 打开 Excel 文件
wb = openpyxl.load_workbook('d:/code/web/src/Amazon商品数据（含图片）.xlsx')
ws = wb.active

print(f"工作表名称: {ws.title}")
print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

# 获取所有图片
images = ws._images
print(f"\n找到 {len(images)} 个图片对象")

# 遍历所有图片
for idx, img in enumerate(images):
    try:
        # 获取图片数据
        if hasattr(img, '_data'):
            image_data = img._data
        elif hasattr(img, 'ref'):
            image_data = img.ref
        else:
            print(f"图片 {idx}: 无法获取数据")
            continue
        
        # 确定图片格式
        if hasattr(img, 'format'):
            img_format = img.format.lower()
        else:
            img_format = 'png'  # 默认格式
        
        # 生成文件名
        filename = f"amazon_product_{idx+1:03d}.{img_format}"
        filepath = os.path.join(output_dir, filename)
        
        # 保存图片
        if isinstance(image_data, bytes):
            with open(filepath, 'wb') as f:
                f.write(image_data)
            print(f"已保存: {filename}")
        else:
            print(f"图片 {idx}: 数据类型不支持")
            
    except Exception as e:
        print(f"处理图片 {idx} 时出错: {e}")

# 也尝试使用 pandas 读取数据来匹配图片位置
df = pd.read_excel('d:/code/web/src/Amazon商品数据（含图片）.xlsx')
print(f"\n数据行数: {len(df)}")
print(f"列名: {list(df.columns)}")

print(f"\n图片已保存到: {output_dir}")
